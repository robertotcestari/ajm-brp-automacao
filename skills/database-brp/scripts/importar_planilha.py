#!/usr/bin/env python3
"""Import the existing BRP control spreadsheet into the local SQLite database."""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl")

sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "lib"))
import brp_db  # noqa: E402


HEADER_TO_FIELD = {
    "n do processo": "numero_processo",
    "no do processo": "numero_processo",
    "numero do processo": "numero_processo",
    "parte contraria": "parte_contraria",
    "tribunal/uf": "tribunal_uf",
    "comarca origem": "comarca_origem",
    "link dos autos": "link_autos",
    "chave": "chave",
    "data de recebimento": "data_recebimento",
    "advogado responsavel": "advogado_responsavel",
    "status da triagem": "status_triagem",
    "audiencia": "audiencia",
    "prazo defesa": "prazo_defesa",
    "prazo de defesa": "prazo_defesa",
    "fonte do prazo": "fonte_prazo",
    "status da pasta": "status_pasta",
    "caminho da pasta": "caminho_pasta",
    "peticao baixada": "peticao_baixada",
    "status da minuta": "status_minuta",
    "teses aplicaveis": "teses_aplicaveis",
}


def norm(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = "".join(c for c in value if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", value).strip().lower()


def cell_value(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", required=True)
    ap.add_argument("--db", default=None)
    args = ap.parse_args()

    planilha = Path(args.planilha)
    if not planilha.exists():
        sys.exit(f"Planilha não encontrada: {planilha}")

    wb = openpyxl.load_workbook(planilha, read_only=True, data_only=True)
    ws = wb.active
    headers = [HEADER_TO_FIELD.get(norm(cell.value)) for cell in ws[1]]

    db_path = brp_db.db_path_from_config(args.db)
    conn = brp_db.connect(db_path)
    inseridos = 0
    atualizados = 0
    pulados = 0
    try:
        brp_db.init_db(conn)
        for row in ws.iter_rows(min_row=2, values_only=True):
            registro = {}
            for field, value in zip(headers, row):
                if field:
                    value = cell_value(value)
                    if value is not None:
                        registro[field] = value
            if not registro.get("numero_processo"):
                pulados += 1
                continue
            result = brp_db.upsert_processo(conn, registro)
            if result["acao"] == "inserido":
                inseridos += 1
            else:
                atualizados += 1
    finally:
        conn.close()
        wb.close()

    print(json.dumps({
        "acao": "planilha_importada",
        "planilha": str(planilha),
        "db": str(db_path),
        "inseridos": inseridos,
        "atualizados": atualizados,
        "pulados_sem_numero": pulados,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
