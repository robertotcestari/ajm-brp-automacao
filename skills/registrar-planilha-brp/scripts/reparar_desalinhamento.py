#!/usr/bin/env python3
"""
Reparo único: corrige o desalinhamento off-by-one que afetou o lote de 01-03/06/2026.

Sintoma: as colunas "automáticas" (Tribunal/UF, Link dos autos, Chave, Data de
recebimento e Caminho da pasta) de cada linha receberam o valor que pertencia à
LINHA DE BAIXO (deslocamento de uma linha). Nº do Processo, Parte, Advogado,
Audiência e Prazo ficaram corretos.

Estratégia (tudo casado pelo Nº do Processo, nunca por posição):
- Tribunal/UF, Link, Chave, Data de recebimento  -> fonte autoritativa = os
  REGISTROS dos scripts batch_registrar*.py / registrar_0139687.py.
- Caminho da pasta / Status da pasta -> recalculado a partir da PRÓPRIA parte+número
  da linha (mesma regra do criar_pasta.py) e validado contra a pasta real no disco.

Por padrão roda em SIMULAÇÃO (mostra o diff). Use --aplicar para gravar (faz backup antes).
"""
import argparse
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

AQUI = Path(__file__).parent
sys.path.insert(0, str(AQUI))
sys.path.insert(0, str(AQUI.parent.parent / "criar-pasta-processo" / "scripts"))

from append_to_planilha import carrega_ou_cria, normaliza_num, vazio  # noqa: E402
from criar_pasta import formata_parte, sanitiza  # noqa: E402

PLANILHA = r"G:\A.Digital\BRP\Gooroo\Planilha de Defesas - Gooroo\Cópia de Defesas BRP 5.xlsx"
BASE = r"G:\A.Digital\BRP"

# colunas reconstruídas pela fonte autoritativa (por número)
COL_DADOS = {
    "tribunal_uf": "Tribunal/UF",
    "link_autos": "Link dos autos",
    "chave": "Chave",
    "data_recebimento": "Data de recebimento",
}


def carrega_autoritativos():
    """Junta os REGISTROS de todos os scripts de lote, indexados pelo número (só dígitos)."""
    modulos = [
        "batch_registrar",
        "batch_registrar_20260605",
        "batch_registrar_20260608",
        "registrar_0139687",
    ]
    auth = {}
    for nome in modulos:
        try:
            mod = __import__(nome)
        except Exception as e:  # pragma: no cover
            print(f"  aviso: não consegui importar {nome}: {e}", file=sys.stderr)
            continue
        for rec in getattr(mod, "REGISTROS", []):
            chave = normaliza_num(rec.get("numero_processo"))
            if chave:
                auth.setdefault(chave, {}).update(rec)
    return auth


def caminho_esperado(parte, numero):
    nome = f"{formata_parte(parte)} - {sanitiza(numero)}"
    return Path(BASE) / nome


_CNJ = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")


def numero_no_caminho(cam):
    """Número CNJ (só dígitos) embutido num caminho de pasta, ou None."""
    if vazio(cam):
        return None
    m = _CNJ.search(str(cam))
    return normaliza_num(m.group(0)) if m else None


def desejado_para_linha(num_cell, parte_cell, cam_cell, auth):
    """Devolve (desejado, tem_fonte, pasta_existe).

    - Tribunal/UF, Link, Chave, Data: só quando há registro autoritativo do número
      (string vazia no registro => limpar a célula deslocada).
    - Caminho da pasta: só é tocado quando está demonstravelmente ERRADO — o número
      embutido no caminho atual difere do número da própria linha (sintoma exato do
      off-by-one) — ou quando está vazio MAS há registro autoritativo (caso da linha
      cujo caminho foi "roubado" pela de cima). Caminhos vazios de linhas sem registro
      NÃO são preenchidos (não fazem parte deste bug).
    - Status da pasta: vira "Criada" quando a pasta da própria linha existe no disco
      e a linha está no escopo (tem registro, ou tinha caminho deslocado/roubado).
    """
    own = normaliza_num(num_cell)
    if not own or vazio(parte_cell):
        return {}, False, False
    rec = auth.get(own)
    desejado = {}

    if rec:
        for json_key, coluna in COL_DADOS.items():
            valor = rec.get(json_key)
            desejado[coluna] = None if vazio(valor) else str(valor).strip()

    parte = parte_cell if not vazio(parte_cell) else (rec or {}).get("parte_contraria", "")
    destino = caminho_esperado(parte, num_cell)
    existe = destino.exists()

    emb = numero_no_caminho(cam_cell)
    caminho_errado = emb is not None and emb != own          # aponta p/ outro processo
    caminho_roubado = vazio(cam_cell) and rec is not None     # vazio mas deveria ter

    if caminho_errado or caminho_roubado:
        desejado["Caminho da pasta"] = str(destino) if existe else None
        if existe:
            desejado["Status da pasta"] = "Criada"
    elif existe and rec is not None:
        # caminho já certo (ou ausente sem ser do bug); garante o status correto
        desejado["Status da pasta"] = "Criada"

    return desejado, bool(rec), existe


def fmt(v):
    if v is None:
        return "(vazio)"
    s = str(v)
    return s if len(s) <= 40 else s[:37] + "..."


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", default=PLANILHA)
    ap.add_argument("--aplicar", action="store_true", help="grava as correções (padrão: só simula)")
    args = ap.parse_args()

    auth = carrega_autoritativos()
    print(f"Fonte autoritativa: {len(auth)} processos.\n")

    wb, ws, colmap = carrega_ou_cria(args.planilha)
    col_proc = colmap["Nº do Processo"]
    col_parte = colmap["Parte Contrária"]

    mudancas = []          # (linha, processo, coluna, antes, depois)
    sem_fonte = []         # linhas corrigidas no caminho mas sem fonte p/ UF/Link/Chave

    for r in range(2, ws.max_row + 1):
        num = ws.cell(row=r, column=col_proc).value
        parte = ws.cell(row=r, column=col_parte).value
        cam = ws.cell(row=r, column=colmap["Caminho da pasta"]).value
        desejado, tem_fonte, _existe = desejado_para_linha(num, parte, cam, auth)
        if not desejado:
            continue
        linha_mudou = False
        for coluna, valor in desejado.items():
            ci = colmap[coluna]
            atual = ws.cell(row=r, column=ci).value
            a = None if vazio(atual) else str(atual).strip()
            d = None if vazio(valor) else str(valor).strip()
            if a != d:
                mudancas.append((r, str(num).strip(), coluna, atual, valor))
                linha_mudou = True
        if linha_mudou and not tem_fonte:
            sem_fonte.append((r, str(num).strip()))

    if not mudancas:
        print("Nada a corrigir — planilha já está alinhada.")
        return

    # agrupa por linha para leitura
    por_linha = {}
    for r, proc, coluna, antes, depois in mudancas:
        por_linha.setdefault((r, proc), []).append((coluna, antes, depois))

    print(f"{len(mudancas)} célula(s) a corrigir em {len(por_linha)} linha(s):\n")
    for (r, proc), itens in sorted(por_linha.items()):
        print(f"  linha {r}  [{proc}]")
        for coluna, antes, depois in itens:
            print(f"      {coluna:<20} {fmt(antes):<42} ->  {fmt(depois)}")

    if sem_fonte:
        print("\n  ATENÇÃO — linhas sem registro autoritativo: corrigi o Caminho/Status,")
        print("  mas Tribunal/UF, Link e Chave NÃO foram verificados (revisar à mão):")
        for r, proc in sem_fonte:
            print(f"      linha {r}  [{proc}]")

    if not args.aplicar:
        print("\n(SIMULAÇÃO — nada gravado. Rode com --aplicar para corrigir.)")
        return

    # backup antes de gravar
    origem = Path(args.planilha)
    carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = origem.with_name(f"{origem.stem}.backup-{carimbo}{origem.suffix}")
    shutil.copy2(origem, backup)
    print(f"\nBackup criado: {backup}")

    for r, proc, coluna, antes, depois in mudancas:
        # IMPORTANTE: ws.cell(..., value=None) NÃO limpa a célula no openpyxl
        # (quando value é None o método ignora). Para esvaziar é preciso setar .value.
        ws.cell(row=r, column=colmap[coluna]).value = depois
    wb.save(args.planilha)
    print(f"OK: {len(mudancas)} célula(s) corrigida(s) em {len(por_linha)} linha(s).")


if __name__ == "__main__":
    main()
