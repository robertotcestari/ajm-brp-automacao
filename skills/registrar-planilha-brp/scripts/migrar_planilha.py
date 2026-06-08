#!/usr/bin/env python3
"""
Migração única: reestrutura a planilha de controle BRP para o schema limpo.

- Mantém TODAS as linhas e os dados já preenchidos (nº, parte, advogado, audiência, prazo).
- Reordena as colunas na ordem lógica do schema (identificação → fluxo → prazos → automação).
- Descarta colunas vazias/perdidas do arquivo antigo.
- Aplica cabeçalho formatado, congela a 1ª linha e cria listas (dropdowns) nos campos de status.

Uso:
  python migrar_planilha.py --origem "Defesas BRP.xlsx" --destino "Defesas BRP - nova.xlsx"
"""
import argparse
import re
import unicodedata
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "Nº do Processo", "Parte Contrária", "Tribunal/UF",
    "Link dos autos", "Chave", "Data de recebimento",
    "Advogado responsável", "Status da triagem",
    "Audiência", "Prazo de defesa", "Fonte do prazo",
    "Status da pasta", "Caminho da pasta", "Petição baixada", "Status da minuta",
    "Teses aplicáveis",
]

# colunas antigas (normalizadas) -> coluna canônica de destino
ORIGEM = {
    "no do processo": "Nº do Processo",
    "n do processo": "Nº do Processo",
    "parte contraria": "Parte Contrária",
    "advogado responsavel": "Advogado responsável",
    "audiencia": "Audiência",
    "prazo defesa": "Prazo de defesa",
    "prazo de defesa": "Prazo de defesa",
}

# listas (dropdowns) por coluna
LISTAS = {
    "Status da triagem": ["Novo", "Em análise", "Distribuído", "Concluído"],
    "Fonte do prazo": ["E-mail", "Autos", "Provisório"],
    "Status da pasta": ["Criada", "Pendente"],
    "Petição baixada": ["Sim", "Não"],
    "Status da minuta": ["Não iniciada", "Rascunho gerado", "Em revisão", "Protocolada"],
}

LARGURAS = {
    "Nº do Processo": 26, "Parte Contrária": 34, "Tribunal/UF": 12,
    "Link dos autos": 30, "Chave": 16, "Data de recebimento": 18, "Advogado responsável": 20,
    "Status da triagem": 16, "Audiência": 18, "Prazo de defesa": 16, "Fonte do prazo": 14,
    "Status da pasta": 14, "Caminho da pasta": 30, "Petição baixada": 14, "Status da minuta": 16,
    "Teses aplicáveis": 26,
}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--origem", required=True)
    ap.add_argument("--destino", required=True)
    args = ap.parse_args()

    src = load_workbook(args.origem)
    ws_src = src.active
    headers_src = [c.value for c in ws_src[1]]
    pos = {}  # canônica -> índice de coluna na origem
    for i, h in enumerate(headers_src, start=1):
        canon = ORIGEM.get(norm(h))
        if canon and canon not in pos:
            pos[canon] = i

    out = Workbook()
    ws = out.active
    ws.title = "Controle BRP"
    arial = Font(name="Arial", size=10)
    arial_bold = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5496")

    # cabeçalho
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = arial_bold
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = LARGURAS.get(h, 16)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    # dados
    linha_saida = 2
    for r in range(2, ws_src.max_row + 1):
        # pula linhas totalmente vazias
        if all(ws_src.cell(row=r, column=c).value in (None, "") for c in range(1, ws_src.max_column + 1)):
            continue
        for j, h in enumerate(HEADERS, start=1):
            if h in pos:
                val = ws_src.cell(row=r, column=pos[h]).value
                cel = ws.cell(row=linha_saida, column=j, value=val)
            else:
                cel = ws.cell(row=linha_saida, column=j)
            cel.font = arial
        linha_saida += 1

    ult = linha_saida - 1
    # dropdowns nos campos de status
    for h, opcoes in LISTAS.items():
        col = HEADERS.index(h) + 1
        letra = ws.cell(row=1, column=col).column_letter
        dv = DataValidation(type="list", formula1='"' + ",".join(opcoes) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letra}2:{letra}{ult}")

    out.save(args.destino)
    print(f"OK: {ult-1} linhas migradas -> {args.destino}")


if __name__ == "__main__":
    main()
