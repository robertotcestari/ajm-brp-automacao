#!/usr/bin/env python3
"""
Anexa ou atualiza uma linha na planilha de controle BRP (.xlsx local).

Regras:
- Dedup pelo número do processo: se já existe, atualiza apenas as células VAZIAS
  (preserva o que a equipe preencheu à mão). Se não existe, insere nova linha.
- Campos sem informação não são inventados; o schema define os padrões de status.

Uso:
  python append_to_planilha.py --planilha "/caminho/Defesas BRP.xlsx" --registro '{...json...}'
  python append_to_planilha.py --planilha "..." --registro-arquivo registro.json
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl --break-system-packages")

# Ordem canônica das colunas (ver assets/schema-planilha.md)
COLUNAS = [
    "Nº do Processo", "Parte Contrária", "Tribunal/UF",
    "Link dos autos", "Chave", "Data de recebimento", "Advogado responsável",
    "Status da triagem", "Audiência", "Prazo de defesa", "Fonte do prazo",
    "Status da pasta", "Caminho da pasta", "Petição baixada", "Status da minuta",
    "Teses aplicáveis",
]

# chave do registro JSON -> nome da coluna
MAPA = {
    "numero_processo": "Nº do Processo",
    "parte_contraria": "Parte Contrária",
    "tribunal_uf": "Tribunal/UF",
    "link_autos": "Link dos autos",
    "chave": "Chave",
    "data_recebimento": "Data de recebimento",
    "advogado_responsavel": "Advogado responsável",
    "status_triagem": "Status da triagem",
    "audiencia": "Audiência",
    "prazo_defesa": "Prazo de defesa",
    "fonte_prazo": "Fonte do prazo",
    "status_pasta": "Status da pasta",
    "caminho_pasta": "Caminho da pasta",
    "peticao_baixada": "Petição baixada",
    "status_minuta": "Status da minuta",
    "teses_aplicaveis": "Teses aplicáveis",
}

# Padrões quando o campo não vem no registro (apenas em INSERÇÃO)
PADROES = {
    "Status da triagem": "Novo",
    "Prazo de defesa": "verificar autos",
    "Status da pasta": "Pendente",
    "Petição baixada": "Não",
    "Status da minuta": "Não iniciada",
}

# Cabeçalhos já existentes na planilha da AJM que correspondem a uma coluna canônica
# (para reaproveitar a coluna em vez de criar uma duplicada). Comparação por forma
# normalizada (sem acento/caixa). Ex.: "prazo defesa" (atual) == "Prazo de defesa".
VARIANTES = {
    "Advogado responsável": ["advogado responsavel"],
    "Audiência": ["audiencia"],
    "Prazo de defesa": ["prazo defesa", "prazo de defesa"],
    "Nº do Processo": ["n do processo", "no do processo", "numero do processo"],
}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def normaliza_num(v):
    return re.sub(r"\D", "", str(v or ""))


def vazio(v):
    return v is None or str(v).strip() == ""


def monta_valores(registro):
    valores = {}
    for chave, coluna in MAPA.items():
        if chave in registro and not vazio(registro[chave]):
            valores[coluna] = str(registro[chave]).strip()
    # mescla observação de audiência no texto da audiência
    obs = registro.get("audiencia_obs")
    if obs and not vazio(obs) and valores.get("Audiência"):
        valores["Audiência"] = f"{valores['Audiência']} — {str(obs).strip()}"
    return valores


def mapeia_colunas(ws):
    """Devolve {coluna_canônica: índice}. Reaproveita colunas já existentes
    (casando por forma normalizada e por variantes conhecidas) e cria as que faltam."""
    existentes = {}
    max_col = 0
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None and str(cell.value).strip():
            existentes[norm(cell.value)] = idx
        max_col = idx
    colmap = {}
    for canon in COLUNAS:
        candidatos = [norm(canon)] + [norm(v) for v in VARIANTES.get(canon, [])]
        achado = next((existentes[c] for c in candidatos if c in existentes), None)
        if achado is None:
            max_col += 1
            ws.cell(row=1, column=max_col, value=canon)
            achado = max_col
            existentes[norm(canon)] = achado
        colmap[canon] = achado
    return colmap


def carrega_ou_cria(caminho):
    p = Path(caminho)
    if p.exists():
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan1"
        for i, c in enumerate(COLUNAS, start=1):
            ws.cell(row=1, column=i, value=c)
    return wb, ws, mapeia_colunas(ws)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", required=True)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--registro")
    g.add_argument("--registro-arquivo")
    args = ap.parse_args()

    if args.registro:
        registro = json.loads(args.registro)
    else:
        registro = json.loads(Path(args.registro_arquivo).read_text(encoding="utf-8"))

    num = registro.get("numero_processo")
    if vazio(num):
        sys.exit("Registro sem 'numero_processo'.")

    valores = monta_valores(registro)

    try:
        wb, ws, colmap = carrega_ou_cria(args.planilha)
    except Exception as e:
        sys.exit(f"Não consegui abrir a planilha ({args.planilha}): {e}. "
                 "Verifique a VPN e se o arquivo não está aberto/bloqueado.")

    col_proc = colmap["Nº do Processo"]
    alvo = normaliza_num(num)

    # procura linha existente
    linha_existente = None
    for r in range(2, ws.max_row + 1):
        if normaliza_num(ws.cell(row=r, column=col_proc).value) == alvo and alvo:
            linha_existente = r
            break

    if linha_existente:
        atualizados = []
        for coluna, valor in valores.items():
            ci = colmap[coluna]
            atual = ws.cell(row=linha_existente, column=ci).value
            if vazio(atual):  # só preenche vazios — preserva edição humana
                ws.cell(row=linha_existente, column=ci, value=valor)
                atualizados.append(coluna)
        wb.save(args.planilha)
        print(json.dumps({"acao": "atualizado", "linha": linha_existente,
                          "processo": num, "campos_preenchidos": atualizados},
                         ensure_ascii=False))
    else:
        nova = ws.max_row + 1
        # aplica padrões nas colunas ainda não definidas
        for coluna, padrao in PADROES.items():
            valores.setdefault(coluna, padrao)
        for coluna, valor in valores.items():
            ws.cell(row=nova, column=colmap[coluna], value=valor)
        wb.save(args.planilha)
        print(json.dumps({"acao": "inserido", "linha": nova, "processo": num},
                         ensure_ascii=False))


if __name__ == "__main__":
    main()
