#!/usr/bin/env python3
"""Backfill: troca os eventos obsoletos "Verificar prazo nos autos" (datados do dia em que a
automação rodou) por eventos de PRAZO DE DEFESA com data calculada.

Para cada evento "Verificar prazo nos autos — <CNJ>" no calendário BRP:
  - procura o processo na planilha de controle pelo número CNJ;
  - se houver Data de recebimento válida → cria/atualiza o evento de prazo de defesa
    (recebimento + 13 dias úteis, sáb/dom pulados, feriados ignorados; PROVISÓRIO) e
    APAGA o evento de verificação;
  - se NÃO houver recebimento → mantém o "verificar" (ainda não dá para calcular).

Idempotente: rodar de novo não duplica (o prazo passa por upsert e o verificar já terá sido
removido). Por padrão SIMULA; use --aplicar para gravar/apagar.

  python backfill_prazos.py            # relatório
  python backfill_prazos.py --aplicar  # executa
"""
import argparse
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

# Credenciais do graph.env (mesmo esquema do batch/dedup)
ENV_FILE = Path(__file__).parents[3] / "config" / "graph.env"
if ENV_FILE.exists():
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

sys.path.insert(0, str(Path(__file__).parent))
from criar_evento_graph import (get_token, acha_calendario, garante_categoria,
                                 monta_evento, upsert_evento, snapshot_eventos,
                                 http, GRAPH, CATEGORIAS, PREFIXO_ASSUNTO, datas)

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl --break-system-packages")

CONFIG = json.loads((Path(__file__).parents[3] / "config" / "brp.config.json")
                    .read_text(encoding="utf-8"))
MAILBOX = os.environ.get("BRP_MAILBOX", CONFIG.get("brp_mailbox"))
CALENDARIO = os.environ.get("BRP_CALENDARIO", CONFIG.get("brp_calendario", "BRP"))
PLANILHA = (
    CONFIG.get("planilha_claude_path")
    or CONFIG.get("planilha_path")
    or CONFIG.get("planilha_original_path")
)
if not PLANILHA:
    sys.exit("Configure planilha_claude_path em config/brp.config.json para usar o backfill.")

CNJ_RE = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def so_num(s):
    return re.sub(r"\D", "", str(s or ""))


def carrega_planilha():
    """Mapa: dígitos do CNJ -> {numero, parte, recebimento}."""
    ws = openpyxl.load_workbook(PLANILHA, data_only=True).active
    col = {}
    for i, c in enumerate(ws[1], 1):
        h = norm(c.value)
        if "processo" in h and "proc" not in col:
            col["proc"] = i
        elif "parte" in h:
            col["parte"] = i
        elif "recebimento" in h:
            col["rec"] = i
    faltando = [k for k in ("proc", "rec") if k not in col]
    if faltando:
        sys.exit(f"Planilha sem colunas esperadas {faltando}. Cabeçalhos: "
                 f"{[c.value for c in ws[1]]}")
    mapa = {}
    for r in range(2, ws.max_row + 1):
        num = ws.cell(row=r, column=col["proc"]).value
        if not num or not so_num(num):
            continue
        mapa[so_num(num)] = {
            "numero": str(num).strip(),
            "parte": str(ws.cell(row=r, column=col.get("parte", 0)).value or "").strip()
                     if "parte" in col else "",
            "recebimento": ws.cell(row=r, column=col["rec"]).value,
        }
    return mapa


def nota_provisoria(recebimento):
    d = datas.parse_data(recebimento)
    return (f"Prazo PROVISÓRIO: recebimento {datas.fmt_br(d)} + {datas.DIAS_UTEIS_DEFESA} "
            f"dias úteis (sem feriados). Confirmar nos autos.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--aplicar", action="store_true",
                    help="Cria os prazos e apaga os 'verificar' (sem isto, apenas simula)")
    args = ap.parse_args()

    plan = carrega_planilha()
    token = get_token()
    cal_id = acha_calendario(token, MAILBOX, CALENDARIO)
    eventos = snapshot_eventos(token, MAILBOX, cal_id)

    prefixo_ver = PREFIXO_ASSUNTO["verificar"]
    converter, sem_rec = [], []
    for ev in eventos:
        subj = ev.get("subject", "") or ""
        if not subj.startswith(prefixo_ver):
            continue
        m = CNJ_RE.search(subj)
        if not m:
            continue
        rec = plan.get(so_num(m.group(0)))
        if rec and datas.parse_data(rec["recebimento"]):
            converter.append((ev, rec))
        else:
            sem_rec.append((ev, m.group(0)))

    print(f"Calendario '{CALENDARIO}': {len(eventos)} eventos.")
    print(f"Eventos 'verificar' encontrados: {len(converter) + len(sem_rec)}")
    print(f"  -> a converter em prazo (tem recebimento): {len(converter)}")
    print(f"  -> mantidos (sem recebimento na planilha):  {len(sem_rec)}\n")

    for ev, rec in converter:
        prazo = datas.fmt_br(datas.prazo_defesa(rec["recebimento"]))
        print(f"  [prazo {prazo}] {rec['numero']} ({rec['parte'][:30]}) "
              f"recebimento {datas.fmt_br(datas.parse_data(rec['recebimento']))} -> apaga verificar")
    if sem_rec:
        print("\n  Mantidos sem recebimento:")
        for _, num in sem_rec:
            print(f"    - {num}")

    if not converter:
        print("\nNada a converter.")
        return
    if not args.aplicar:
        print(f"\n(SIMULACAO) Rode com --aplicar para criar {len(converter)} prazos e "
              f"apagar {len(converter)} eventos 'verificar'.")
        return

    garante_categoria(token, MAILBOX, *CATEGORIAS["prazo"])
    criados, apagados, erros = 0, 0, 0
    print("\nExecutando...")
    for ev, rec in converter:
        try:
            inicio = datas.fmt_iso(datas.prazo_defesa(rec["recebimento"]))
            evento = monta_evento("prazo", rec["numero"], rec["parte"], inicio, 0,
                                  nota_provisoria(rec["recebimento"]))
            acao, _ = upsert_evento(token, MAILBOX, cal_id, "prazo", rec["numero"], evento)
            http("DELETE", f"{GRAPH}/users/{MAILBOX}/calendars/{cal_id}/events/{ev['id']}", token)
            criados += 1
            apagados += 1
            print(f"  [+] prazo {acao} / verificar apagado — {rec['numero']}")
        except SystemExit as e:
            print(f"  [!] ERRO {rec['numero']}: {e}", file=sys.stderr)
            erros += 1
    print(f"\nPrazos lançados: {criados}, verificar apagados: {apagados}, erros: {erros}")


if __name__ == "__main__":
    main()
